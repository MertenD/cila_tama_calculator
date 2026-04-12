"""
Berechnet die TAMA-Flächen (Total Abdominal Muscle Area) für alle Patienten.

TAMA ist eine Messgröße für die Muskelfläche im Bauchbereich.
Sie wird berechnet als: Gesamte Muskulatur MINUS Innere Organe = Rumpfmuskulatur

Das Skript durchsucht alle Patientenordner und berechnet die TAMA-Fläche für jeden
verfügbaren Zeitpunkt (Baseline und Follow-Ups).
"""

import csv
import os
import glob
import SimpleITK as sitk
import numpy as np

# Neu: Excel-Lesen für Körpergrößen
from typing import Optional, Dict
# Neu für Parallelisierung
from typing import Any
import re

try:
    import openpyxl
except Exception:  # openpyxl ist optional, App soll auch ohne Excel laufen
    openpyxl = None

# Pfad zum Hauptordner mit den Patientendaten
rootPath = "sources/nrrd/Sarkome_Phyton/2025-12-10"

# Optional: Excel-Datei mit Patienten-Stammdaten
DEFAULT_PATIENT_XLSX = os.path.join("sources", "NK_Doktorarbeit_Liste_1.3.xlsx")
DEFAULT_PATIENT_SHEET = "Patienten"
DEFAULT_PATIENT_ID_COL = "ID"
DEFAULT_PATIENT_HEIGHT_COL = "Körpergrößen"
DEFAULT_PATIENT_SEX_COL = "Geschlecht"

# Hounsfield-Einheiten-Schwellenwerte: Definieren den Dichtebereich von Muskelgewebe in CT-Bildern
# Nur Pixel mit Werten zwischen -29 und 150 werden als Muskel erkannt
muscleHU = (-29., 150.)

# SMI (Skeletal Muscle Index) Schwellenwerte in cm^2 / m^2
# Quelle/Definition: vom Projekt vorgegeben
SMI_THRESHOLDS_CM2_PER_M2 = {
    "M": 55.0,
    "F": 39.0,
}


def normalize_patient_id(value) -> Optional[str]:
    """Normalisiert Patienten-IDs für das Matching Excel <-> Ordnername.

    - Entfernt alles außer Ziffern
    - Entfernt führende Nullen
    - Gibt eine kanonische Ziffern-String-Repräsentation zurück (ohne führende Nullen)

    Beispiele:
      "0003088882" -> "3088882"
      3088882      -> "3088882"
      "3088882"   -> "3088882"
    """
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    digits = re.sub(r"\D+", "", s)
    if not digits:
        return None

    # int() entfernt führende Nullen robust; bei extrem langen IDs fallback auf lstrip
    try:
        return str(int(digits))
    except Exception:
        stripped = digits.lstrip('0')
        return stripped or "0"


def parse_height_m(value) -> Optional[float]:
    """Parst Körpergröße zu Metern (float).

    Akzeptiert z.B. "1,76", "1.76", "176", "176 cm".
    Heuristik:
      - Werte < 10 -> als Meter
      - Werte >= 10 -> als Zentimeter
    """
    if value is None:
        return None

    s = str(value).strip()
    if not s or s.lower() in {"n/a", "na", "nan", "none"}:
        return None

    # Entferne Einheiten/Spaces, toleriert Komma als Dezimaltrenner
    s = s.replace(" ", "").replace("cm", "").replace("m", "")
    s = s.replace(",", ".")

    # Nur eine Zahl extrahieren
    m = re.search(r"\d+(?:\.\d+)?", s)
    if not m:
        return None

    try:
        num = float(m.group(0))
    except ValueError:
        return None

    if num <= 0:
        return None

    # Heuristik: 1.76m vs 176cm
    if num < 10:
        return round(num, 3)
    return round(num / 100.0, 3)


def parse_sex(value) -> Optional[str]:
    """Parst Geschlecht aus Excel.

    Erwartet typischerweise "F", "M" oder "" (leer).
    Gibt "F"/"M" zurück oder None.
    """
    if value is None:
        return None
    s = str(value).strip().upper()
    if not s:
        return None
    if s in {"F", "M"}:
        return s
    return None


def load_patient_metadata(
        xlsx_path: str = DEFAULT_PATIENT_XLSX,
        sheet_name: str = DEFAULT_PATIENT_SHEET,
        id_col_name: str = DEFAULT_PATIENT_ID_COL,
        height_col_name: str = DEFAULT_PATIENT_HEIGHT_COL,
        sex_col_name: str = DEFAULT_PATIENT_SEX_COL,
) -> Dict[str, dict]:
    """Lädt Patienten-Metadaten aus Excel.

    Rückgabe: Mapping normalized_pat_id -> {"height_m": Optional[float], "sex": Optional[str]}
    """
    if openpyxl is None:
        return {}

    if not xlsx_path or not os.path.exists(xlsx_path):
        return {}

    meta: Dict[str, dict] = {}

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return {}

        ws = wb[sheet_name]

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return {}

        header_index = {str(h).strip(): idx for idx, h in enumerate(header) if h is not None}
        if id_col_name not in header_index:
            return {}

        id_idx = header_index[id_col_name]
        h_idx = header_index.get(height_col_name)
        s_idx = header_index.get(sex_col_name)

        for r in rows:
            if not r:
                continue

            raw_id = r[id_idx] if id_idx < len(r) else None
            norm_id = normalize_patient_id(raw_id)
            if not norm_id:
                continue

            raw_h = r[h_idx] if (h_idx is not None and h_idx < len(r)) else None
            raw_s = r[s_idx] if (s_idx is not None and s_idx < len(r)) else None

            height_m = parse_height_m(raw_h)
            sex = parse_sex(raw_s)

            # Wir nehmen den Eintrag auch dann mit, wenn nur eines von beidem vorhanden ist
            if height_m is None and sex is None:
                continue

            meta[norm_id] = {"height_m": height_m, "sex": sex}

    except Exception:
        return {}

    return meta


def load_patient_heights(
        xlsx_path: str = DEFAULT_PATIENT_XLSX,
        sheet_name: str = DEFAULT_PATIENT_SHEET,
        id_col_name: str = DEFAULT_PATIENT_ID_COL,
        height_col_name: str = DEFAULT_PATIENT_HEIGHT_COL
) -> Dict[str, float]:
    """Lädt Körpergrößen aus der Excel und liefert Mapping: normalized_pat_id -> height_m.

    (Kompatibilitäts-Wrapper; intern wird `load_patient_metadata()` genutzt.)
    """
    meta = load_patient_metadata(
        xlsx_path=xlsx_path,
        sheet_name=sheet_name,
        id_col_name=id_col_name,
        height_col_name=height_col_name,
    )
    return {k: v["height_m"] for k, v in meta.items() if v.get("height_m") is not None}


def pad_patient_id_like_folder(folder_patient_id: str, excel_norm_id: Optional[str]) -> Optional[str]:
    """Padde eine normalisierte Excel-ID (ohne führende Nullen) auf die Länge des Ordnernamens.

    Damit bleibt die Ausgabe-Spalte `patId` weiterhin identisch zum Ordnernamen.
    """
    if excel_norm_id is None:
        return None

    if not folder_patient_id:
        return excel_norm_id

    digits = re.sub(r"\D+", "", str(folder_patient_id))
    if not digits:
        return excel_norm_id

    return str(excel_norm_id).zfill(len(digits))


def scaleToOriginal(orgImage, segmentation):
    """
    Passt die Auflösung einer Markierung an das Original-Bild an.

    Die Segmentierungsdaten (Markierungen von Muskeln und Organen) können eine andere
    Auflösung haben als das Original-CT-Bild. Diese Funktion bringt beide auf dieselbe
    Auflösung, damit sie miteinander verrechnet werden können.

    Args:
        orgImage: Das Original-CT-Bild mit seiner ursprünglichen Auflösung
        segmentation: Die Markierung (z.B. Muskel-Markierung), die angepasst werden soll

    Returns:
        Die Markierung in der Auflösung des Original-Bildes
    """
    res = sitk.Resample(segmentation, orgImage.GetSize(), interpolator=sitk.sitkNearestNeighbor,
                        outputOrigin=orgImage.GetOrigin(), outputSpacing=orgImage.GetSpacing(),
                        outputDirection=orgImage.GetDirection(), outputPixelType=sitk.sitkUInt8)
    return res


# Zusatz-Segmentierungen (rechte Muskeln) – robuste Pattern (case-insensitive)
RIGHT_MUSCLE_SEGMENT_REGEX = {
    # Beispiele aus den Daten: "M_ quadratus lumborum rechts", "M_ erector spinae rechts", "M_ psoas major rechts"
    "quadratusLumborumRight": re.compile(r"quadratus\s*lumborum\s*rechts", re.IGNORECASE),
    "erectorSpinaeRight": re.compile(r"erector\s*spinae\s*rechts", re.IGNORECASE),
    "psoasMajorRight": re.compile(r"psoas\s*major\s*rechts", re.IGNORECASE),
}


def find_nrrd_file_by_regex(directory: str, filename_regex: "re.Pattern") -> Optional[str]:
    """Findet die erste .nrrd-Datei, deren Dateiname auf `filename_regex` matcht (ohne .json)."""
    if not directory or not os.path.exists(directory):
        return None

    try:
        files = glob.glob(os.path.join(directory, "*.nrrd"))
    except Exception:
        return None

    files = [f for f in files if not f.endswith('.json')]
    for f in files:
        name = os.path.basename(f)
        try:
            if filename_regex.search(name):
                return f
        except Exception:
            continue
    return None


def calculate_muscle_area_from_seg(
        org_img: "sitk.Image",
        seg_img: "sitk.Image",
        dcm_thresholded: "sitk.Image",
) -> float:
    """Berechnet Muskel-Fläche (mm²) aus einer Segmentierung, gefiltert mit HU-Threshold."""
    seg_scaled = scaleToOriginal(org_img, seg_img)
    seg_filtered = sitk.And(seg_scaled, dcm_thresholded)
    return float(getAreaFromSeg(seg_filtered))


def find_nrrd_file(directory, pattern):
    """
    Sucht eine Datei mit einem bestimmten Textmuster im Dateinamen.

    Die Markierungsdateien haben spezifische Namen (z.B. enthält "Musc" für Muskeln
    oder "Visc" für innere Organe). Diese Funktion findet die richtige Datei.

    Args:
        directory: Ordner, in dem gesucht werden soll
        pattern: Textmuster nach dem gesucht wird (z.B. "Musc" oder "Visc")

    Returns:
        Pfad zur gefundenen Datei oder None, falls keine Datei gefunden wurde
    """
    search_pattern = os.path.join(directory, f"*{pattern}*.nrrd")
    files = glob.glob(search_pattern)
    # Filtere JSON-Dateien aus (wir brauchen nur die eigentlichen Bilddateien)
    files = [f for f in files if not f.endswith('.json')]

    if files:
        return files[0]
    return None


def find_dicom_series(nrrd_directory):
    """
    Findet und lädt die Original-CT-Bildserie im Verzeichnis.

    CT-Scans bestehen aus vielen einzelnen Bildern (Schichten), die zusammen
    einen 3D-Scan ergeben. Diese Funktion lädt alle Bildschichten und setzt sie
    zu einem vollständigen 3D-Bild zusammen.

    Args:
        nrrd_directory: Verzeichnis, in dem nach den Original-CT-Daten gesucht wird

    Returns:
        Das zusammengesetzte 3D-CT-Bild oder None, falls keine Daten gefunden wurden
    """
    # Finde DICOM-Unterordner (erkennbar an einem Punkt im Namen)
    dicom_dirs = [d for d in os.listdir(nrrd_directory)
                  if os.path.isdir(os.path.join(nrrd_directory, d)) and '.' in d]

    if not dicom_dirs:
        return None

    dicom_dir = os.path.join(nrrd_directory, dicom_dirs[0])

    # Lade alle CT-Bildschichten und setze sie zu einem 3D-Bild zusammen
    try:
        reader = sitk.ImageSeriesReader()
        dicom_names = reader.GetGDCMSeriesFileNames(dicom_dir)

        if not dicom_names:
            return None

        reader.SetFileNames(dicom_names)
        image = reader.Execute()
        return image
    except Exception as e:
        print(f"    Fehler beim Laden der CT-Bildserie: {e}")
        return None


def get_date_from_dicom(nrrd_directory):
    """
    Extrahiert das Aufnahmedatum des CT-Scans aus den Bilddaten.

    Jedes CT-Bild enthält Metadaten (zusätzliche Informationen), unter anderem
    das Datum, an dem der Scan durchgeführt wurde. Diese Funktion liest dieses Datum aus.

    Args:
        nrrd_directory: Verzeichnis mit den CT-Daten

    Returns:
        Datum als String im Format YYYY-MM-DD (z.B. "2025-01-15") oder "Unknown"
    """
    # Finde CT-Daten-Unterordner
    dicom_dirs = [d for d in os.listdir(nrrd_directory)
                  if os.path.isdir(os.path.join(nrrd_directory, d)) and '.' in d]

    if not dicom_dirs:
        return "Unknown"

    dicom_dir = os.path.join(nrrd_directory, dicom_dirs[0])

    # Finde erste CT-Bilddatei
    dcm_files = glob.glob(os.path.join(dicom_dir, "*.dcm"))

    if not dcm_files:
        return "Unknown"

    try:
        # Lese CT-Bilddatei und ihre Metadaten
        reader = sitk.ImageFileReader()
        reader.SetFileName(dcm_files[0])
        reader.LoadPrivateTagsOn()
        reader.ReadImageInformation()

        # Versuche verschiedene Datumsfelder aus den Metadaten zu lesen
        # (StudyDate, SeriesDate oder AcquisitionDate)
        date_str = None

        if reader.HasMetaDataKey("0008|0020"):  # StudyDate
            date_str = reader.GetMetaData("0008|0020")
        elif reader.HasMetaDataKey("0008|0021"):  # SeriesDate
            date_str = reader.GetMetaData("0008|0021")
        elif reader.HasMetaDataKey("0008|0022"):  # AcquisitionDate
            date_str = reader.GetMetaData("0008|0022")

        if date_str and len(date_str) == 8:  # Format: YYYYMMDD
            # Formatiere das Datum mit Bindestrichen: YYYY-MM-DD
            return f"{date_str[0:4]}-{date_str[4:6]}-{date_str[6:8]}"

    except Exception as e:
        print(f"    Warnung: Konnte Datum nicht aus den CT-Daten extrahieren: {e}")

    return "Unknown"


def getAreaFromSeg(segmentation):
    """
    Berechnet die Fläche einer Markierung in Quadratmillimetern.

    Zählt alle markierten Bildpunkte und multipliziert mit der realen Größe eines
    Pixels, um die tatsächliche Fläche in mm² zu erhalten.

    Args:
        segmentation: Die Markierung, deren Fläche berechnet werden soll

    Returns:
        Fläche in mm² (Quadratmillimeter)
    """
    # Zähle alle markierten Bildpunkte (Pixel mit Wert ungleich 0)
    pixelCount = np.count_nonzero(sitk.GetArrayFromImage(segmentation))

    # Berechne die reale Fläche: Anzahl der Pixel × Pixelbreite × Pixelhöhe
    area = pixelCount * segmentation.GetSpacing()[0] * segmentation.GetSpacing()[1]
    return area


def calculate_smi(tama_area_mm2: Optional[float], height_m: Optional[float]) -> Optional[float]:
    """Berechnet SMI = (TAMA_cm2) / (height_m^2).

    TAMA liegt im Code in mm² vor; für SMI wird in cm² / m² gerechnet.
    => cm² = mm² / 100
    """
    if tama_area_mm2 is None or height_m is None:
        return None
    try:
        h = float(height_m)
        if h <= 0:
            return None
        tama_cm2 = float(tama_area_mm2) / 100.0
        return round(tama_cm2 / (h ** 2), 4)
    except Exception:
        return None


def classify_muscle_wasting(sex: Optional[str], smi: Optional[float]) -> Optional[bool]:
    """DEPRECATED: Muskelschwund-Flags werden nicht mehr ausgegeben.

    (Funktion bleibt zur Abwärtskompatibilität im Code; wird nicht mehr genutzt.)
    """
    if smi is None:
        return None
    if not sex:
        return None
    key = str(sex).strip().upper()
    if key not in SMI_THRESHOLDS_CM2_PER_M2:
        return None
    return smi < SMI_THRESHOLDS_CM2_PER_M2[key]


def process_timepoint(
        nrrd_dir,
        patient_id,
        timepoint,
        patient_heights: Optional[Dict[str, float]] = None,
        patient_metadata: Optional[Dict[str, dict]] = None
):
    """
    Berechnet die TAMA-Fläche für einen bestimmten Untersuchungszeitpunkt eines Patienten.

    Lädt die Muskel- und Organmarkierungen, das Original-CT-Bild, berechnet die
    Rumpfmuskulatur (TAMA) und gibt die Fläche zurück.

    Args:
        nrrd_dir: Pfad zum Ordner mit den Bilddaten
        patient_id: Eindeutige Nummer des Patienten
        timepoint: Zeitpunkt (z.B. "Baseline" oder "Follow-Up 1")
        patient_heights: Optionales Mapping normalized_pat_id -> height_m (legacy)
        patient_metadata: Optionales Mapping normalized_pat_id -> {"height_m": float|None, "sex": "F"|"M"|None}

    Returns:
        Dictionary mit Ergebnissen oder None bei Fehler
    """
    print(f"  Verarbeite {patient_id} - {timepoint}")

    # Suche nach den Dateien mit den Muskel- und Organmarkierungen
    musc_file = find_nrrd_file(nrrd_dir, "Musc")
    visc_file = find_nrrd_file(nrrd_dir, "Visc")
    vert_file = find_nrrd_file(nrrd_dir, "Vert")

    # Zusätzliche rechte Muskel-Segmentierungen (optional)
    ql_right_file = find_nrrd_file_by_regex(nrrd_dir, RIGHT_MUSCLE_SEGMENT_REGEX["quadratusLumborumRight"])
    es_right_file = find_nrrd_file_by_regex(nrrd_dir, RIGHT_MUSCLE_SEGMENT_REGEX["erectorSpinaeRight"])
    pm_right_file = find_nrrd_file_by_regex(nrrd_dir, RIGHT_MUSCLE_SEGMENT_REGEX["psoasMajorRight"])

    if not musc_file:
        print(f"    ⚠ Keine Muskel-Markierungsdatei gefunden")
        return None

    if not visc_file:
        print(f"    ⚠ Keine Organ-Markierungsdatei gefunden")
        return None

    try:
        # Lade das Original-CT-Bild (besteht aus mehreren Schichten)
        org_img = find_dicom_series(nrrd_dir)

        if org_img is None:
            print(f"    ⚠ Keine CT-Bildserie gefunden")
            return None

        # HU-Threshold einmal berechnen (wird für alle Muskel-Teilflächen wiederverwendet)
        dcm_thresholded = sitk.BinaryThreshold(
            org_img,
            lowerThreshold=muscleHU[0],
            upperThreshold=muscleHU[1],
            outsideValue=0,
            insideValue=1
        )

        # Lade die Markierungen für Muskeln und Organe
        musc_img = sitk.ReadImage(musc_file, sitk.sitkUInt8)
        visc_img = sitk.ReadImage(visc_file, sitk.sitkUInt8)
        vert_img = sitk.ReadImage(vert_file, sitk.sitkUInt8)

        # Bringe die Markierungen auf die gleiche Auflösung wie das Original-CT-Bild
        musc_scaled = scaleToOriginal(org_img, musc_img)
        visc_scaled = scaleToOriginal(org_img, visc_img)
        vert_scaled = scaleToOriginal(org_img, vert_img)

        # Berechne TAMA = Muskel MINUS Organe
        # (Ziehe die Organflächen von der Muskelfläche ab, um nur Rumpfmuskulatur zu erhalten)
        tama_img = sitk.And(musc_scaled, sitk.Not(visc_scaled))
        # Berechne TAMA mit zusätzlicher Subtraktion der Wirbelsäule (Vert), um nur die reine Rumpfmuskulatur zu erhalten
        tama_img_vert_subtracted = sitk.And(tama_img, sitk.Not(vert_scaled))

        # Kombiniere die anatomische Markierung mit der Dichtefilterung
        # Nur Bereiche, die SOWOHL als Muskel markiert SIND als auch die richtige Dichte HABEN, bleiben übrig
        tama_filtered = sitk.And(tama_img, dcm_thresholded)
        tama_filtered_vert_subtracted = sitk.And(tama_img_vert_subtracted, dcm_thresholded)

        # Berechne die Fläche der gefilterten TAMA-Markierung
        tama_area_mm2 = getAreaFromSeg(tama_filtered)
        tama_area_mm2_vert_subtracted = getAreaFromSeg(tama_filtered_vert_subtracted)

        # Rechte Teilmuskeln (optional; fehlende Dateien -> None)
        ql_right_area_mm2 = None
        es_right_area_mm2 = None
        pm_right_area_mm2 = None

        def _try_area(file_path: Optional[str]) -> Optional[float]:
            if not file_path:
                return None
            try:
                seg = sitk.ReadImage(file_path, sitk.sitkUInt8)
                return round(calculate_muscle_area_from_seg(org_img, seg, dcm_thresholded), 2)
            except Exception as e:
                print(f"    Warnung: Konnte Segmentierung nicht verarbeiten ({os.path.basename(file_path)}): {e}")
                return None

        ql_right_area_mm2 = _try_area(ql_right_file)
        es_right_area_mm2 = _try_area(es_right_file)
        pm_right_area_mm2 = _try_area(pm_right_file)

        # Lese das Datum des CT-Scans aus
        scan_date = get_date_from_dicom(nrrd_dir)

        print(f"    ✓ TAMA-Fläche: {tama_area_mm2:.0f} mm²")
        print(f"    ✓ TAMA-Fläche Vert Subtrahiert: {tama_area_mm2_vert_subtracted:.0f} mm²")
        if ql_right_area_mm2 is not None:
            print(f"    ✓ M. quadratus lumborum rechts: {ql_right_area_mm2:.0f} mm²")
        if es_right_area_mm2 is not None:
            print(f"    ✓ M. erector spinae rechts: {es_right_area_mm2:.0f} mm²")
        if pm_right_area_mm2 is not None:
            print(f"    ✓ M. psoas major rechts: {pm_right_area_mm2:.0f} mm²")

        # Optional: Körpergröße/Geschlecht aus Excel ergänzen
        height_m = None
        sex = None
        norm_id = None

        if patient_metadata:
            norm_id = normalize_patient_id(patient_id)
            if norm_id and norm_id in patient_metadata:
                md = patient_metadata.get(norm_id) or {}
                height_m = md.get("height_m")
                sex = md.get("sex")

        # Fallback: nur Körpergrößen (alte API)
        if height_m is None and patient_heights:
            norm_id = norm_id or normalize_patient_id(patient_id)
            if norm_id and norm_id in patient_heights:
                height_m = patient_heights.get(norm_id)

        # SMI (für beide TAMA-Varianten)
        smi_vert_not_sub = calculate_smi(tama_area_mm2, height_m)
        smi_vert_sub = calculate_smi(tama_area_mm2_vert_subtracted, height_m)

        return {
            "patId": patient_id,
            "timepoint": timepoint,
            "date": scan_date,
            "bodyHeightM": height_m,
            "sex": sex,
            "tamaAreaVertNotSubtracted": round(tama_area_mm2, 2),
            "tamaAreaVertSubtracted": round(tama_area_mm2_vert_subtracted, 2),
            "smiVertNotSubtracted": smi_vert_not_sub,
            "smiVertSubtracted": smi_vert_sub,
            # Neu: rechte Muskel-Teilflächen (mm²)
            "areaQuadratusLumborumRight": ql_right_area_mm2,
            "areaErectorSpinaeRight": es_right_area_mm2,
            "areaPsoasMajorRight": pm_right_area_mm2,
        }

    except Exception as e:
        print(f"    ✗ Fehler bei Verarbeitung: {e}")
        return None


def _process_timepoint_task(args: Dict[str, Any]) -> Optional[dict]:
    """Worker-Wrapper für Parallelisierung (ProcessPool-friendly).

    Erwartet ein Dict mit Keys:
      - nrrd_path, patient_id, timepoint
      - patient_metadata (optional), patient_heights (optional)
      - hu_min, hu_max (optional)
      - sitk_threads (optional)

    Rückgabe: result-dict oder None.
    """
    # Oversubscription vermeiden: pro Prozess interne Threads reduzieren.
    # SimpleITK unterstützt das je nach Build.
    sitk_threads = args.get("sitk_threads")
    if sitk_threads is not None:
        try:
            sitk.ProcessObject.SetGlobalDefaultNumberOfThreads(int(sitk_threads))
        except Exception:
            pass

    # HU-Range pro Task setzen (globale Variable wird nur im Worker-Prozess verändert)
    hu_min = args.get("hu_min")
    hu_max = args.get("hu_max")
    if hu_min is not None and hu_max is not None:
        try:
            global muscleHU
            muscleHU = (float(hu_min), float(hu_max))
        except Exception:
            pass

    return process_timepoint(
        args["nrrd_path"],
        args["patient_id"],
        args["timepoint"],
        patient_heights=args.get("patient_heights"),
        patient_metadata=args.get("patient_metadata"),
    )


def main():
    """Hauptfunktion: Durchsucht alle Patienten und berechnet TAMA-Flächen für alle Zeitpunkte."""

    print("=" * 70)
    print("TAMA-Flächen Berechnung")
    print("=" * 70)
    print()

    # Liste für alle Ergebnisse
    results = []
    patient_dirs = []

    # Prüfe, ob der Hauptordner existiert
    if not os.path.exists(rootPath):
        print(f"❌ Fehler: Pfad nicht gefunden: {rootPath}")
        return

    # Durchsuche alle Patientenordner
    for patient_id in os.listdir(rootPath):
        patient_path = os.path.join(rootPath, patient_id)

        # Überspringe, wenn es keine Ordner sind
        if not os.path.isdir(patient_path):
            continue

        # Die erwartete Ordnerstruktur: PatID/Befundverlauf 1/0/[Baseline|Follow-Up X]/nrrd
        befund_path = os.path.join(patient_path, "Befundverlauf 1", "0")

        if not os.path.exists(befund_path):
            print(f"⚠ Überspringe {patient_id} - keine Standard-Struktur")
            continue

        # Durchsuche alle Zeitpunkte (Baseline und Follow-Ups) für diesen Patienten
        for timepoint_name in os.listdir(befund_path):
            timepoint_path = os.path.join(befund_path, timepoint_name)

            if not os.path.isdir(timepoint_path):
                continue

            nrrd_path = os.path.join(timepoint_path, "nrrd")

            # Prüfe, ob der nrrd-Ordner mit den Bilddaten existiert
            if not os.path.exists(nrrd_path):
                continue

            # Extrahiere die Nummer aus dem Zeitpunkt-Namen für die Sortierung
            # Baseline erhält die Nummer 0, Follow-Up X erhält die Nummer X
            if "Baseline" in timepoint_name:
                sort_key = 0
            else:
                import re
                match = re.search(r'\d+', timepoint_name)
                sort_key = int(match.group()) if match else 999

            # Speichere die Informationen für diesen Zeitpunkt
            patient_dirs.append({
                "nrrd_path": nrrd_path,
                "patient_id": patient_id,
                "timepoint": timepoint_name,
                "sort_key": sort_key
            })

    # Sortiere die Zeitpunkte nach Patient-ID und Zeitpunkt-Nummer
    patient_dirs.sort(key=lambda x: (x['patient_id'], x['sort_key']))

    print(f"Gefunden: {len(patient_dirs)} Zeitpunkte für {len(set(pd['patient_id'] for pd in patient_dirs))} Patienten")
    print()

    # Lade Körpergrößen/Geschlecht (optional)
    patient_metadata = load_patient_metadata()
    patient_heights = {k: v["height_m"] for k, v in patient_metadata.items() if v.get("height_m") is not None}

    if patient_metadata:
        print(f"✓ Patienten-Metadaten geladen für {len(patient_metadata)} Patienten (aus Excel)")
    else:
        print("ℹ Keine Patienten-Metadaten geladen (Excel fehlt oder nicht lesbar) — fahre ohne fort")

    # Verarbeite jeden gefundenen Zeitpunkt
    for i, info in enumerate(patient_dirs, 1):
        print(f"[{i}/{len(patient_dirs)}] Patient {info['patient_id']} - {info['timepoint']}")

        result = process_timepoint(
            info['nrrd_path'],
            info['patient_id'],
            info['timepoint'],
            patient_heights=patient_heights,
            patient_metadata=patient_metadata
        )

        if result:
            results.append(result)

        print()

    # Speichere alle Ergebnisse in einer CSV-Datei
    if results:
        output_file = "tama_areas.csv"

        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            fieldnames = [
                "patId",
                "timepoint",
                "date",
                "bodyHeightM",
                "sex",
                "tamaAreaVertSubtracted",
                "tamaAreaVertNotSubtracted",
                "smiVertSubtracted",
                "smiVertNotSubtracted",
                # Neu
                "areaQuadratusLumborumRight",
                "areaErectorSpinaeRight",
                "areaPsoasMajorRight",
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            writer.writerows(results)

        print("=" * 70)
        print(f"✓ Ergebnisse gespeichert in: {output_file}")
        print(f"✓ Verarbeitet: {len(results)} von {len(patient_dirs)} Zeitpunkten")
        print("=" * 70)

        # Zeige beispielhaft die Ergebnisse für Patient 3088882
        print("\nErgebnisse für Patient 3088882:")
        patient_results = [r for r in results if r['patId'] == '0003088882']
        patient_results.sort(key=lambda x: x['timepoint'])

        for r in patient_results:
            print(f"  {r['timepoint']:15s}: {r['tamaAreaVertSubtracted']:8.0f} mm² ({r['date']})")
            print(f"  {r['timepoint']:15s}: {r['tamaAreaVertNotSubtracted']:8.0f} mm² ({r['date']})")

    else:
        print("❌ Keine Ergebnisse zum Speichern")


if __name__ == "__main__":
    main()
